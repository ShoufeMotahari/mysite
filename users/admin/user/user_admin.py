import io
import logging
import os
import platform

from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.core.exceptions import ValidationError
from django.db.models import Count
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import path
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from django_jalali.admin.filters import JDateFieldListFilter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.services.email_service.email_service import EmailValidator, EmailService
from filemanager.models import User
from users.admin.filter.email_status_filter import EmailStatusFilter
from users.admin.filter.user_type_filter import UserTypeFilter
from users.admin.comment_admin import HasCommentsFilter, CommentInline
from users.forms.email_form import EmailForm, QuickEmailTemplateForm
from users.models.admin_message.admin_message import AdminMessage
from users.models.user.user_type import UserType

logger = logging.getLogger(__name__)


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    # List display with user types
    list_display = [
        "username",
        "image_tag",
        "email",
        "mobile",
        "full_name_display",
        "user_type_display",
        "is_active",
        "is_staff",
        "email_status",
        "phone_status",
        "comments_count",
        "last_login",
        "created_at",
    ]

    list_editable = ["is_active", "is_staff"]

    list_filter = [
        "is_active",
        "is_staff",
        "is_superuser",
        UserTypeFilter,
        "is_phone_verified",
        "is_email_verified",
        HasCommentsFilter,
        EmailStatusFilter,
        ("created_at", JDateFieldListFilter),
        ("last_login", JDateFieldListFilter),
    ]

    search_fields = [
        "username",
        "email",
        "mobile",
        "first_name",
        "last_name",
        "slug",
        "comments__content",
        "user_type__name",
    ]

    # Enhanced actions - including export actions
    actions = [
        "send_email_action",
        "activate_users",
        "deactivate_users",
        "verify_emails",
        "verify_phones",
        "change_user_type_action",
        "promote_to_staff",
        "demote_from_staff",
        "export_to_excel",
        "export_to_pdf",
    ]

    # Pagination
    list_per_page = 50
    list_max_show_all = 200

    # Ordering
    ordering = ["-created_at"]

    # Add inlines
    inlines = [CommentInline]

    # Enhanced fieldsets for editing existing users
    fieldsets = (
        ("اطلاعات کاربری", {"fields": ("username", "email", "mobile", "slug")}),
        (
            "تصویر کاربر",
            {
                "fields": ("image",),
            },
        ),
        ("اطلاعات شخصی", {"fields": ("first_name", "last_name", "bio", "birth_date")}),
        (
            "نوع کاربری و دسترسی",
            {
                "fields": ("user_type", "is_active", "is_staff", "is_superuser"),
                "description": "نوع کاربری تعیین کننده سطح دسترسی کاربر است",
            },
        ),
        ("وضعیت تایید", {"fields": ("is_phone_verified", "is_email_verified")}),
        (
            "مجوزها",
            {"fields": ("groups", "user_permissions"), "classes": ("collapse",)},
        ),
        (
            "آمار فعالیت",
            {
                "fields": ("posts_count", "comments_count", "last_activity"),
                "classes": ("collapse",),
            },
        ),
        (
            "تاریخ‌ها",
            {
                "fields": ("last_login", "date_joined", "created_at"),
                "classes": ("collapse",),
            },
        ),
    )

    # Add fieldsets specifically for adding new users
    add_fieldsets = (
        (
            "اطلاعات کاربری ضروری",
            {
                "classes": ("wide",),
                "fields": ("username", "email", "mobile", "password1", "password2"),
            },
        ),
        (
            "اطلاعات شخصی",
            {
                "classes": ("wide",),
                "fields": ("first_name", "last_name"),
            },
        ),
        (
            "نوع کاربری و دسترسی",
            {
                "classes": ("wide",),
                "fields": ("user_type", "is_active", "is_staff"),
                "description": "نوع کاربری تعیین کننده سطح دسترسی کاربر است",
            },
        ),
        (
            "تصویر کاربر",
            {
                "classes": ("wide",),
                "fields": ("image",),
            },
        ),
    )

    readonly_fields = [
        "created_at",
        "image_preview",
        "date_joined",
        "last_login",
        "last_activity",
        "posts_count",
        "comments_count",
    ]

    def get_persian_font_path(self):
        """Get available Persian font path from system"""
        system = platform.system()

        if system == "Windows":
            paths = [
                "C:/Windows/Fonts/tahoma.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/calibri.ttf",
            ]
        elif system == "Linux":
            paths = [
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/TTF/arial.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            ]
        elif system == "Darwin":  # macOS
            paths = [
                "/System/Library/Fonts/Arial.ttf",
                "/Library/Fonts/Arial.ttf",
                "/System/Library/Fonts/Helvetica.ttc",
            ]
        else:
            return None

        for path in paths:
            if os.path.exists(path):
                return path
        return None

    def export_to_excel(self, request, queryset):
        import openpyxl

        """Export selected users to Persian Excel file"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "کاربران"

            # Set RTL direction for the worksheet
            ws.sheet_view.rightToLeft = True

            # Define headers in Persian
            headers = [
                "نام کاربری",
                "نام کامل",
                "ایمیل",
                "موبایل",
                "نوع کاربری",
                "وضعیت فعال",
                "عضو کادر",
                "ایمیل تایید شده",
                "تلفن تایید شده",
                "تعداد نظرات",
                "آخرین ورود",
                "تاریخ عضویت",
            ]

            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Write data
            for row, user in enumerate(queryset, 2):
                data = [
                    user.username or "",
                    user.get_full_name() or "بدون نام",
                    user.email or "",
                    user.mobile or "",
                    user.user_type.name if user.user_type else "پیش‌فرض",
                    "فعال" if user.is_active else "غیرفعال",
                    "بله" if user.is_staff else "خیر",
                    "تایید شده" if user.is_email_verified else "تایید نشده",
                    "تایید شده" if user.is_phone_verified else "تایید نشده",
                    user.comments.count(),
                    (
                        user.last_login.strftime("%Y/%m/%d %H:%M")
                        if user.last_login
                        else "هرگز"
                    ),
                    (
                        user.created_at.strftime("%Y/%m/%d %H:%M")
                        if user.created_at
                        else ""
                    ),
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Prepare response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.xlsx"

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Log success
            logger.info(f"✅ Excel export completed by admin {request.user.username}")
            logger.info(f"   📊 Exported {queryset.count()} users to {filename}")

            self.message_user(
                request,
                f"فایل اکسل شامل {queryset.count()} کاربر با موفقیت ایجاد شد.",
                level=messages.SUCCESS,
            )

            return response

        except Exception as e:
            logger.error(
                f"❌ Excel export failed by admin {request.user.username}: {str(e)}"
            )
            self.message_user(
                request, f"خطا در ایجاد فایل اکسل: {str(e)}", level=messages.ERROR
            )
            return redirect("admin:users_user_changelist")

    export_to_excel.short_description = "دریافت خروجی اکسل از کاربران انتخاب شده"

    def export_to_pdf(self, request, queryset):
        from reportlab.lib import colors

        """Export selected users to PDF with proper Persian BIDI support"""
        try:
            # Import BIDI support libraries
            try:
                from arabic_reshaper import reshape
                from bidi.algorithm import get_display

                bidi_available = True
            except ImportError:
                bidi_available = False
                logger.warning(
                    "BIDI libraries not available. Install: pip install python-bidi arabic-reshaper"
                )

            # Create response
            response = HttpResponse(content_type="application/pdf")
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Create PDF document
            doc = SimpleDocTemplate(
                response,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20,
            )

            # Font registration with better Persian support
            persian_font_available = False
            registered_font_name = "Helvetica"

            # Try to get Persian font
            font_path = self.get_persian_font_path()
            if font_path:
                try:
                    pdfmetrics.registerFont(TTFont("SystemFont", font_path))
                    registered_font_name = "SystemFont"
                    persian_font_available = True
                    logger.info(f"Successfully registered font from: {font_path}")
                except Exception as font_error:
                    logger.warning(f"Failed to register system font: {font_error}")

            # Helper function to process Persian text
            def process_persian_text(text):
                """Process Persian text for proper RTL display"""
                if not text or not isinstance(text, str):
                    return str(text) if text else ""

                if not bidi_available:
                    return text

                try:
                    # First reshape Arabic/Persian characters
                    reshaped_text = reshape(text)
                    # Then apply BIDI algorithm
                    bidi_text = get_display(reshaped_text)
                    return bidi_text
                except Exception as e:
                    logger.warning(f"Failed to process Persian text '{text}': {e}")
                    return text

            # Build content
            story = []
            styles = getSampleStyleSheet()

            # Create custom styles with RTL support
            if persian_font_available:
                title_style = ParagraphStyle(
                    "CustomTitle",
                    parent=styles["Title"],
                    fontName=registered_font_name,
                    fontSize=16,
                    alignment=1,  # Center
                    spaceAfter=20,
                    textColor=colors.HexColor("#366092"),
                    # Add RTL support if available
                    wordWrap="RTL" if bidi_available else "LTR",
                )

                info_style = ParagraphStyle(
                    "CustomInfo",
                    parent=styles["Normal"],
                    fontName=registered_font_name,
                    fontSize=11,
                    alignment=2,  # Right align for RTL
                    spaceAfter=6,
                    wordWrap="RTL" if bidi_available else "LTR",
                )
            else:
                title_style = styles["Title"]
                info_style = styles["Normal"]

            # Add title and metadata with proper Persian processing
            if persian_font_available:
                title_text = process_persian_text("گزارش کاربران سیستم")
                story.append(Paragraph(title_text, title_style))
                story.append(Spacer(1, 15))

                current_time = timezone.now().strftime("%Y/%m/%d - %H:%M")
                admin_name = request.user.get_full_name() or request.user.username

                # Process entire strings with BIDI, not just parts
                if bidi_available:
                    # Create complete Persian sentences first, then process
                    admin_line = f"{admin_name} :تهیه شده توسط"  # Reverse order for RTL
                    date_line = f"{current_time} :تاریخ تهیه"
                    count_line = f"{queryset.count()} :تعداد کاربران"

                    # Process complete lines
                    admin_line_processed = process_persian_text(admin_line)
                    date_line_processed = process_persian_text(date_line)
                    count_line_processed = process_persian_text(count_line)

                    story.append(Paragraph(admin_line_processed, info_style))
                    story.append(Paragraph(date_line_processed, info_style))
                    story.append(Paragraph(count_line_processed, info_style))
                else:
                    # Fallback without BIDI - use simple approach
                    story.append(Paragraph(f"تهیه شده توسط: {admin_name}", info_style))
                    story.append(Paragraph(f"تاریخ تهیه: {current_time}", info_style))
                    story.append(
                        Paragraph(f"تعداد کاربران: {queryset.count()}", info_style)
                    )
            else:
                # Fallback to English
                story.append(Paragraph("Users Export Report", title_style))
                story.append(Spacer(1, 15))

                current_time = timezone.now().strftime("%Y-%m-%d %H:%M")
                admin_name = request.user.get_full_name() or request.user.username

                story.append(Paragraph(f"Generated by: {admin_name}", info_style))
                story.append(Paragraph(f"Export date: {current_time}", info_style))
                story.append(Paragraph(f"Total users: {queryset.count()}", info_style))

            story.append(Spacer(1, 20))

            # Prepare table data with Persian processing
            table_data = []

            # Headers with proper Persian processing
            if persian_font_available and bidi_available:
                headers = [
                    process_persian_text("نام کاربری"),
                    process_persian_text("نام کامل"),
                    process_persian_text("ایمیل"),
                    process_persian_text("موبایل"),
                    process_persian_text("نوع کاربری"),
                    process_persian_text("وضعیت"),
                    process_persian_text("تاریخ عضویت"),
                ]
            elif persian_font_available:
                headers = [
                    "نام کاربری",
                    "نام کامل",
                    "ایمیل",
                    "موبایل",
                    "نوع کاربری",
                    "وضعیت",
                    "تاریخ عضویت",
                ]
            else:
                headers = [
                    "Username",
                    "Full Name",
                    "Email",
                    "Mobile",
                    "User Type",
                    "Status",
                    "Join Date",
                ]

            table_data.append(headers)

            # Data rows with proper Persian processing
            for user in queryset:
                if persian_font_available and bidi_available:
                    # Process all Persian text properly
                    username = user.username or "-"
                    full_name = process_persian_text(user.get_full_name() or "بدون نام")
                    email = user.email or "-"
                    mobile = user.mobile or "-"
                    user_type = process_persian_text(
                        user.user_type.name if user.user_type else "پیش‌فرض"
                    )
                    status = process_persian_text(
                        "فعال" if user.is_active else "غیرفعال"
                    )
                    join_date = (
                        user.created_at.strftime("%Y/%m/%d") if user.created_at else "-"
                    )

                    row = [
                        username,
                        full_name,
                        email,
                        mobile,
                        user_type,
                        status,
                        join_date,
                    ]

                elif persian_font_available:
                    # Without BIDI processing
                    row = [
                        user.username or "-",
                        user.get_full_name() or "بدون نام",
                        user.email or "-",
                        user.mobile or "-",
                        user.user_type.name if user.user_type else "پیش‌فرض",
                        "فعال" if user.is_active else "غیرفعال",
                        (
                            user.created_at.strftime("%Y/%m/%d")
                            if user.created_at
                            else "-"
                        ),
                    ]
                else:
                    # English fallback
                    full_name = user.get_full_name()
                    if full_name:
                        full_name = str(full_name)
                    else:
                        full_name = "No name"

                    user_type = (
                        str(user.user_type.name) if user.user_type else "Default"
                    )

                    row = [
                        user.username or "-",
                        full_name,
                        user.email or "-",
                        user.mobile or "-",
                        user_type,
                        "Active" if user.is_active else "Inactive",
                        (
                            user.created_at.strftime("%Y-%m-%d")
                            if user.created_at
                            else "-"
                        ),
                    ]

                table_data.append(row)

            # Create table with appropriate widths
            col_widths = [
                1.2 * inch,
                1.8 * inch,
                2.2 * inch,
                1.3 * inch,
                1.5 * inch,
                1 * inch,
                1.2 * inch,
            ]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Enhanced table styling for RTL
            table_style = [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    (
                        registered_font_name
                        if persian_font_available
                        else "Helvetica-Bold"
                    ),
                ),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                # Data rows styling
                (
                    "FONTNAME",
                    (0, 1),
                    (-1, -1),
                    registered_font_name if persian_font_available else "Helvetica",
                ),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                # Alignment for Persian columns (right-aligned)
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "RIGHT" if persian_font_available and bidi_available else "CENTER",
                ),
                # Grid and borders
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#366092")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]

            # Add alternating row colors
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table_style.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.lightgrey)
                    )
                else:
                    table_style.append(("BACKGROUND", (0, i), (-1, i), colors.white))

            table.setStyle(TableStyle(table_style))
            story.append(table)

            # Build PDF
            doc.build(story)

            # Enhanced logging
            logger.info(f"✅ PDF export completed by admin {request.user.username}")
            logger.info(f"   📊 Exported {queryset.count()} users to {filename}")
            logger.info(
                f"   🔤 Font used: {registered_font_name}, Persian support: {persian_font_available}"
            )
            logger.info(f"   📄 BIDI support: {bidi_available}")

            success_msg = f"فایل PDF شامل {queryset.count()} کاربر با موفقیت ایجاد شد."
            if not persian_font_available:
                success_msg += " (از فونت پیش‌فرض استفاده شد)"
            elif not bidi_available:
                success_msg += " (بدون پشتیبانی کامل RTL)"

            self.message_user(request, success_msg, level=messages.SUCCESS)
            return response

        except Exception as e:
            logger.error(
                f"❌ PDF export failed by admin {request.user.username}: {str(e)}"
            )
            logger.exception("PDF export exception details:")
            self.message_user(
                request, f"خطا در ایجاد فایل PDF: {str(e)}", level=messages.ERROR
            )
            return redirect("admin:users_user_changelist")

        export_to_pdf.short_description = "دریافت خروجی PDF از کاربران انتخاب شده"

    def changelist_view(self, request, extra_context=None):
        """Override changelist view to add notifications and export buttons"""
        extra_context = extra_context or {}

        # Add your existing context
        extra_context["show_send_all_email_button"] = True
        extra_context["show_create_user_with_type_button"] = True
        extra_context["show_export_buttons"] = True
        extra_context["user_types"] = UserType.objects.filter(is_active=True)

        # Add admin message notifications
        if request.user.is_staff:
            try:
                unread_messages = AdminMessage.objects.filter(status="unread").order_by(
                    "-created_at"
                )[:3]
                if unread_messages.exists():
                    extra_context["admin_notifications"] = unread_messages
                    extra_context["notification_count"] = unread_messages.count()

                    # Add a Django message for immediate visibility
                    messages.info(
                        request,
                        f"📨 شما {unread_messages.count()} پیام خوانده نشده دارید.",
                        extra_tags="admin-notification",
                    )
            except Exception:
                pass

        return super().changelist_view(request, extra_context=extra_context)

    # Optimize queries
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        # Annotate with comment count for efficiency
        queryset = queryset.annotate(
            comments_count_num=Count("comments"),
            approved_comments_count=Count(
                "comments", filter=Q(comments__is_approved=True)
            ),
        )
        # Select related for efficiency
        queryset = queryset.select_related("user_type")
        return queryset

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height: 100px; border-radius: 5px;" />',
                obj.image.url,
            )
        return "بدون تصویر"

    image_preview.short_description = "پیش‌نمایش تصویر"

    def image_tag(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:40px; border-radius:4px;" />',
                obj.image.url,
            )
        return "-"

    image_tag.short_description = "تصویر"

    def full_name_display(self, obj):
        """Display full name or fallback"""
        full_name = obj.get_full_name().strip()
        if full_name:
            return full_name
        return format_html('<em style="color: #888;">بدون نام</em>')

    full_name_display.short_description = "نام کامل"

    def user_type_display(self, obj):
        """Display user type with color coding"""
        if not obj.user_type:
            return format_html('<em style="color: #888;">نامشخص</em>')

        # Color code based on permissions
        if obj.user_type.can_access_admin:
            color = "darkred"
            weight = "bold"
        elif obj.user_type.can_manage_users:
            color = "purple"
            weight = "bold"
        elif obj.user_type.can_create_content:
            color = "green"
            weight = "normal"
        else:
            color = "blue"
            weight = "normal"

        return format_html(
            '<span style="color: {}; font-weight: {};">{}</span>',
            color,
            weight,
            obj.user_type.name,
        )

    user_type_display.short_description = "نوع کاربری"

    def email_status(self, obj):
        """Display email validation status with colors"""
        if not obj.email:
            return format_html(
                '<span style="color: #888; font-style: italic;">بدون ایمیل</span>'
            )

        status_parts = []

        # Email verification status
        if obj.is_email_verified:
            status_parts.append('<span style="color: green;">✓ تایید شده</span>')
        else:
            status_parts.append('<span style="color: orange;">⚠ تایید نشده</span>')

        # Email validity
        try:
            EmailValidator.validate_email(obj.email)
            status_parts.append('<span style="color: green;">معتبر</span>')
        except ValidationError:
            status_parts.append('<span style="color: red;">نامعتبر</span>')

        return format_html(" | ".join(status_parts))

    email_status.short_description = "وضعیت ایمیل"

    def phone_status(self, obj):
        """Display phone verification status"""
        if not obj.mobile:
            return format_html(
                '<span style="color: #888; font-style: italic;">بدون شماره</span>'
            )

        if obj.is_phone_verified:
            return format_html('<span style="color: green;">✓ تایید شده</span>')
        else:
            return format_html('<span style="color: orange;">⚠ تایید نشده</span>')

    phone_status.short_description = "وضعیت تلفن"

    def comments_count(self, obj):
        """Display comment count with breakdown"""
        total = getattr(obj, "comments_count_num", obj.comments.count())
        approved = getattr(
            obj,
            "approved_comments_count",
            obj.comments.filter(is_approved=True).count(),
        )

        if total == 0:
            return format_html('<span style="color: #888;">بدون نظر</span>')

        pending = total - approved
        parts = [f"<strong>{total}</strong> کل"]

        if approved > 0:
            parts.append(f'<span style="color: green;">{approved} تایید</span>')

        if pending > 0:
            parts.append(f'<span style="color: orange;">{pending} انتظار</span>')

        return format_html(" | ".join(parts))

    comments_count.short_description = "نظرات"

    # Enhanced actions
    def send_email_action(self, request, queryset):
        selected = queryset.values_list("id", flat=True)
        selected_count = len(selected)

        logger.info(
            f"👨‍💼 Admin {request.user.username} initiated email action for {selected_count} users"
        )
        logger.debug(f"Selected user IDs: {list(selected)}")

        request.session["selected_users"] = list(selected)
        return HttpResponseRedirect(reverse("admin:send_email"))

    send_email_action.short_description = "ارسال ایمیل به کاربران انتخاب شده"

    def activate_users(self, request, queryset):
        """Activate selected users"""
        updated = queryset.update(is_active=True)
        logger.info(f"👨‍💼 Admin {request.user.username} activated {updated} users")
        self.message_user(request, f"{updated} کاربر فعال شد.", level=messages.SUCCESS)

    activate_users.short_description = "فعال کردن کاربران انتخاب شده"

    def deactivate_users(self, request, queryset):
        """Deactivate selected users"""
        updated = queryset.update(is_active=False)
        logger.info(f"👨‍💼 Admin {request.user.username} deactivated {updated} users")
        self.message_user(
            request, f"{updated} کاربر غیرفعال شد.", level=messages.WARNING
        )

    deactivate_users.short_description = "غیرفعال کردن کاربران انتخاب شده"

    def verify_emails(self, request, queryset):
        """Verify emails for selected users"""
        updated = (
            queryset.filter(email__isnull=False)
            .exclude(email="")
            .update(is_email_verified=True)
        )
        logger.info(
            f"👨‍💼 Admin {request.user.username} verified emails for {updated} users"
        )
        self.message_user(
            request, f"ایمیل {updated} کاربر تایید شد.", level=messages.SUCCESS
        )

    verify_emails.short_description = "تایید ایمیل کاربران انتخاب شده"

    def verify_phones(self, request, queryset):
        """Verify phones for selected users"""
        updated = (
            queryset.filter(mobile__isnull=False)
            .exclude(mobile="")
            .update(is_phone_verified=True)
        )
        logger.info(
            f"👨‍💼 Admin {request.user.username} verified phones for {updated} users"
        )
        self.message_user(
            request, f"شماره تلفن {updated} کاربر تایید شد.", level=messages.SUCCESS
        )

    verify_phones.short_description = "تایید شماره تلفن کاربران انتخاب شده"

    def change_user_type_action(self, request, queryset):
        """Change user type for selected users"""
        selected = queryset.values_list("id", flat=True)
        request.session["selected_users_for_type_change"] = list(selected)
        return HttpResponseRedirect(reverse("admin:change_user_type"))

    change_user_type_action.short_description = "تغییر نوع کاربری کاربران انتخاب شده"

    def promote_to_staff(self, request, queryset):
        """Promote users to staff status"""
        updated = queryset.update(is_staff=True)
        logger.info(
            f"👨‍💼 Admin {request.user.username} promoted {updated} users to staff"
        )
        self.message_user(
            request, f"{updated} کاربر به عضو کادر ارتقا یافت.", level=messages.SUCCESS
        )

    promote_to_staff.short_description = "ارتقا به عضو کادر"

    def demote_from_staff(self, request, queryset):
        """Demote users from staff status"""
        updated = queryset.update(is_staff=False)
        logger.info(
            f"👨‍💼 Admin {request.user.username} demoted {updated} users from staff"
        )
        self.message_user(
            request, f"{updated} کاربر از عضویت کادر خارج شد.", level=messages.WARNING
        )

    demote_from_staff.short_description = "خروج از عضویت کادر"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("send-email/", self.send_email_view, name="send_email"),
            path("send-email-all/", self.send_email_all_view, name="send_email_all"),
            path(
                "change-user-type/", self.change_user_type_view, name="change_user_type"
            ),
            path(
                "create-user-with-type/",
                self.create_user_with_type_view,
                name="create_user_with_type",
            ),
            # Export URLs
            path(
                "export-all-excel/", self.export_all_excel_view, name="export_all_excel"
            ),
            path("export-all-pdf/", self.export_all_pdf_view, name="export_all_pdf"),
        ]
        return custom_urls + urls

    def export_all_excel_view(self, request):
        """Export all users to Excel"""
        all_users = User.objects.all()
        return self.export_to_excel(request, all_users)

    def export_all_pdf_view(self, request):
        """Export all users to PDF"""
        all_users = User.objects.all()
        return self.export_to_pdf(request, all_users)

    def create_user_with_type_view(self, request):
        """Create user with specific type"""
        if request.method == "POST":
            return self._handle_create_user_post(request)
        else:
            return self._handle_create_user_get(request)

    def _handle_create_user_get(self, request):
        """Handle GET request for create user form"""
        user_types = UserType.objects.filter(is_active=True)
        context = {
            "title": "ایجاد کاربر با نوع مشخص",
            "user_types": user_types,
        }
        return render(request, "admin/create_user_with_type.html", context)

    def _handle_create_user_post(self, request):
        """Handle POST request for create user"""
        try:
            # Get form data
            username = request.POST.get("username")
            email = request.POST.get("email")
            mobile = request.POST.get("mobile")
            first_name = request.POST.get("first_name", "")
            last_name = request.POST.get("last_name", "")
            user_type_id = request.POST.get("user_type")
            password = request.POST.get("password")
            is_active = request.POST.get("is_active") == "on"
            is_staff = request.POST.get("is_staff") == "on"

            # Validation
            if not any([username, email, mobile]):
                messages.error(
                    request,
                    "حداقل یکی از فیلدهای نام کاربری، ایمیل یا موبایل باید پر شود.",
                )
                return self._handle_create_user_get(request)

            if not password:
                messages.error(request, "رمز عبور الزامی است.")
                return self._handle_create_user_get(request)

            # Get user type
            user_type = None
            if user_type_id:
                try:
                    user_type = UserType.objects.get(id=user_type_id, is_active=True)
                except UserType.DoesNotExist:
                    messages.error(request, "نوع کاربری انتخاب شده معتبر نیست.")
                    return self._handle_create_user_get(request)

            # Create user
            user = User.objects.create_user(
                username=username or None,
                email=email or None,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_active=is_active,
                is_staff=is_staff,
            )

            # Set additional fields
            if mobile:
                user.mobile = mobile
            if user_type:
                user.user_type = user_type

            user.save()

            logger.info(
                f"👨‍💼 Admin {request.user.username} created user: {user.get_display_name()} with type: {user_type.name if user_type else 'None'}"
            )

            messages.success(
                request,
                f'کاربر {user.get_display_name()} با نوع کاربری "{user_type.name if user_type else "پیش‌فرض"}" ایجاد شد.',
            )
            return redirect("admin:users_user_changelist")

        except Exception as e:
            logger.error(
                f"❌ Error creating user by admin {request.user.username}: {str(e)}"
            )
            messages.error(request, f"خطا در ایجاد کاربر: {str(e)}")
            return self._handle_create_user_get(request)

    def change_user_type_view(self, request):
        """Change user type for selected users"""
        if request.method == "POST":
            return self._handle_change_type_post(request)
        else:
            return self._handle_change_type_get(request)

    def _handle_change_type_get(self, request):
        """Handle GET request for change type form"""
        selected_users = request.session.get("selected_users_for_type_change", [])
        if not selected_users:
            messages.error(request, "هیچ کاربری انتخاب نشده است.")
            return redirect("admin:users_user_changelist")

        users = User.objects.filter(id__in=selected_users)
        user_types = UserType.objects.filter(is_active=True)

        context = {
            "title": "تغییر نوع کاربری",
            "users": users,
            "user_types": user_types,
            "total_users": len(users),
        }
        return render(request, "admin/change_user_type.html", context)

    def _handle_change_type_post(self, request):
        """Handle POST request for change type"""
        try:
            selected_users = request.session.get("selected_users_for_type_change", [])
            new_type_id = request.POST.get("user_type")

            if not selected_users:
                messages.error(request, "هیچ کاربری انتخاب نشده است.")
                return redirect("admin:users_user_changelist")

            # Get new user type
            new_type = None
            if new_type_id:
                try:
                    new_type = UserType.objects.get(id=new_type_id, is_active=True)
                except UserType.DoesNotExist:
                    messages.error(request, "نوع کاربری انتخاب شده معتبر نیست.")
                    return self._handle_change_type_get(request)

            # Update users
            users = User.objects.filter(id__in=selected_users)
            updated_count = 0

            for user in users:
                old_type = user.user_type
                user.user_type = new_type

                # Auto-update staff status based on new type
                if new_type and new_type.can_access_admin:
                    user.is_staff = True
                elif (
                        old_type
                        and old_type.can_access_admin
                        and new_type
                        and not new_type.can_access_admin
                ):
                    user.is_staff = False

                user.save()
                updated_count += 1

                logger.info(
                    f"👨‍💼 Admin {request.user.username} changed user type for {user.get_display_name()} from {old_type.name if old_type else 'None'} to {new_type.name if new_type else 'None'}"
                )

            # Clear session
            if "selected_users_for_type_change" in request.session:
                del request.session["selected_users_for_type_change"]

            type_name = new_type.name if new_type else "پیش‌فرض"
            messages.success(
                request, f'نوع کاربری {updated_count} کاربر به "{type_name}" تغییر کرد.'
            )
            return redirect("admin:users_user_changelist")

        except Exception as e:
            logger.error(
                f"❌ Error changing user types by admin {request.user.username}: {str(e)}"
            )
            messages.error(request, f"خطا در تغییر نوع کاربری: {str(e)}")
            return self._handle_change_type_get(request)

    # Email management methods
    def send_email_all_view(self, request):
        """Send email to all users view"""
        logger.info(
            f"👨‍💼 Admin {request.user.username} accessing send email to all users"
        )

        # Get all user IDs
        all_user_ids = list(User.objects.values_list("id", flat=True))
        request.session["selected_users"] = all_user_ids
        request.session["email_all_users"] = True

        logger.info(f"📊 Preparing to send email to all {len(all_user_ids)} users")

        return HttpResponseRedirect(reverse("admin:send_email"))

    def send_email_view(self, request):
        """Enhanced send email view with direct composition"""
        if request.method == "POST":
            return self._handle_email_post(request)
        else:
            return self._handle_email_get(request)

    def _handle_email_post(self, request):
        """Handle POST request for email sending"""
        form = EmailForm(request.POST)

        # Handle quick template loading via AJAX
        if request.POST.get('action') == 'load_template':
            template_type = request.POST.get('template_type')
            if template_type and template_type != 'custom':
                template_content = QuickEmailTemplateForm.get_template_content(template_type)
                return JsonResponse({
                    'success': True,
                    'subject': template_content['subject'],
                    'content': template_content['content']
                })
            return JsonResponse({'success': False})

        if not form.is_valid():
            logger.warning(f"⚠️ Invalid form submission by {request.user.username}")
            logger.debug(f"Form errors: {form.errors}")
            return self._render_email_form(request, form)

        try:
            # Get selected users
            selected_users = request.session.get("selected_users", [])
            is_email_all = request.session.get("email_all_users", False)

            if not selected_users:
                logger.warning(
                    f"⚠️ No users selected for email sending by {request.user.username}"
                )
                messages.error(request, "هیچ کاربری انتخاب نشده است.")
                return redirect("admin:users_user_changelist")

            logger.info(
                f"🎯 Processing email form submission by {request.user.username}"
            )
            logger.info(f"Selected users count: {len(selected_users)}")
            logger.info(f"Email all users: {is_email_all}")
            logger.info(f"Subject: {form.cleaned_data.get('subject', 'None')}")

            # Get recipient users
            users = User.objects.filter(id__in=selected_users)

            # Use EmailService directly
            email_service = EmailService()
            success, message, details = email_service.send_email(
                recipients=users,
                subject=form.cleaned_data['subject'],
                content=form.cleaned_data['content'],
                sender_info=f"Admin: {request.user.username}"
            )

            # Process results
            self._process_email_results(request, [(success, message, details)], is_email_all)

            # Clear session
            if "selected_users" in request.session:
                del request.session["selected_users"]
            if "email_all_users" in request.session:
                del request.session["email_all_users"]

            return redirect("admin:users_user_changelist")

        except Exception as e:
            error_msg = f"خطا در ارسال ایمیل: {str(e)}"
            messages.error(request, error_msg)
            logger.error(
                f"❌ Exception in admin email sending by {request.user.username}: {str(e)}"
            )
            logger.exception("Full exception details:")
            return self._render_email_form(request, form)

    def _handle_email_get(self, request):
        """Handle GET request for email form display"""
        form = EmailForm()
        template_form = QuickEmailTemplateForm()
        logger.info(f"📋 Email form displayed to admin {request.user.username}")
        return self._render_email_form(request, form, template_form)

    def _render_email_form(self, request, form, template_form=None):
        """Render the email form with validation info"""
        selected_users = request.session.get("selected_users", [])
        is_email_all = request.session.get("email_all_users", False)
        users = User.objects.filter(id__in=selected_users)

        # Add validation info to context
        valid_users, invalid_users = EmailValidator.validate_users(users)

        logger.debug(
            f"Form display: {len(valid_users)} valid, {len(invalid_users)} invalid users"
        )

        context = {
            "form": form,
            "template_form": template_form or QuickEmailTemplateForm(),
            "users": users,
            "valid_users": valid_users,
            "invalid_users": invalid_users,
            "is_email_all": is_email_all,
            "total_users": len(users),
            "title": "ارسال ایمیل به همه کاربران" if is_email_all else "ارسال ایمیل",
        }

        return render(request, "admin/send_email.html", context)

    def _process_email_results(self, request, results, is_email_all=False):
        """Process email sending results and display appropriate messages"""
        if results and results[0][0]:  # Check if successful
            success, message, details = results[0]

            # Create detailed success message
            total_users = details.get("total_users", 0)
            valid_users = details.get("valid_users", 0)
            invalid_users = details.get("invalid_users", 0)