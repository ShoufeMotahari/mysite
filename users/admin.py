# users/admin.py - Enhanced User Admin Panel with User Types
import io
import logging
import os
import platform

# Import django-jalali admin components
from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.core.exceptions import ValidationError
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django_jalali.admin.filters import JDateFieldListFilter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Import forms and services
from users.forms.forms import EmailForm
from users.managers.email_manager import EmailManager, SendEmailCommand

# Import models
from users.models import (
    AdminMessage,
    AdminMessageReadStatus,
    AdminMessageReply,
    Comment,
    PasswordEntry,
    User,
    UserType,
)
from users.services.email_service import EmailValidator

from .models import EmailTemplate

# Setup logging
logger = logging.getLogger("users")
password_logger = logging.getLogger(__name__)


from django.contrib import admin, messages
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html



class UserTypeFilter(admin.SimpleListFilter):
    """Filter users by user type"""

    title = "نوع کاربری"
    parameter_name = "user_type"

    def lookups(self, request, model_admin):
        user_types = UserType.objects.filter(is_active=True)
        return [(ut.id, ut.name) for ut in user_types] + [("none", "بدون نوع")]

    def queryset(self, request, queryset):
        if self.value() == "none":
            return queryset.filter(user_type__isnull=True)
        elif self.value():
            return queryset.filter(user_type_id=self.value())
        return queryset


class HasCommentsFilter(admin.SimpleListFilter):
    """Custom filter to show users with or without comments"""

    title = "وضعیت نظرات"
    parameter_name = "has_comments"

    def lookups(self, request, model_admin):
        return (
            ("yes", "دارای نظر"),
            ("no", "بدون نظر"),
            ("approved", "دارای نظر تایید شده"),
            ("pending", "دارای نظر در انتظار تایید"),
        )

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(comments__isnull=False).distinct()
        elif self.value() == "no":
            return queryset.filter(comments__isnull=True).distinct()
        elif self.value() == "approved":
            return queryset.filter(comments__is_approved=True).distinct()
        elif self.value() == "pending":
            return queryset.filter(comments__is_approved=False).distinct()
        return queryset


class EmailStatusFilter(admin.SimpleListFilter):
    """Filter users by email validation status"""

    title = "وضعیت ایمیل"
    parameter_name = "email_status"

    def lookups(self, request, model_admin):
        return (
            ("valid", "ایمیل معتبر"),
            ("invalid", "ایمیل نامعتبر"),
            ("no_email", "بدون ایمیل"),
            ("verified", "ایمیل تایید شده"),
            ("unverified", "ایمیل تایید نشده"),
        )

    def queryset(self, request, queryset):
        if self.value() == "valid":
            return queryset.exclude(email__isnull=True).exclude(email="")
        elif self.value() == "invalid":
            return queryset.filter(Q(email__isnull=True) | Q(email=""))
        elif self.value() == "no_email":
            return queryset.filter(Q(email__isnull=True) | Q(email=""))
        elif self.value() == "verified":
            return queryset.filter(is_email_verified=True)
        elif self.value() == "unverified":
            return queryset.filter(is_email_verified=False)
        return queryset


class CommentInline(admin.TabularInline):
    """Inline for user comments"""

    model = Comment
    extra = 0
    readonly_fields = ["created_at", "updated_at"]
    fields = ["content", "is_approved", "is_active", "created_at", "updated_at"]

    def get_queryset(self, request):
        """Only show comments related to this user"""
        return super().get_queryset(request).select_related("user")


@admin.register(UserType)
class UserTypeAdmin(admin.ModelAdmin):
    """Admin for User Types"""

    list_display = [
        "name",
        "slug",
        "permissions_summary",
        "limits_summary",
        "users_count",
        "is_active",
        "is_default",
        "created_at",
    ]
    list_editable = ["is_active"]
    list_filter = ["is_active", "is_default", "can_access_admin", "created_at"]
    search_fields = ["name", "slug", "description"]
    readonly_fields = ["created_at", "updated_at"]
    prepopulated_fields = {"slug": ("name",)}

    fieldsets = (
        (
            "اطلاعات پایه",
            {"fields": ("name", "slug", "description", "is_active", "is_default")},
        ),
        (
            "مجوزهای محتوا",
            {
                "fields": (
                    "can_create_content",
                    "can_edit_content",
                    "can_delete_content",
                )
            },
        ),
        (
            "مجوزهای مدیریتی",
            {"fields": ("can_manage_users", "can_view_analytics", "can_access_admin")},
        ),
        (
            "محدودیت‌های محتوا",
            {
                "fields": (
                    "max_posts_per_day",
                    "max_comments_per_day",
                    "max_file_upload_size_mb",
                )
            },
        ),
        ("تاریخ‌ها", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )

    actions = [
        "activate_user_types",
        "deactivate_user_types",
        "create_predefined_types",
    ]

    def get_queryset(self, request):
        """Annotate with user count"""
        return super().get_queryset(request).annotate(users_count_num=Count("user"))

    def permissions_summary(self, obj):
        """Show a summary of permissions"""
        perms = []
        if obj.can_create_content:
            perms.append('<span style="color: green;">ایجاد</span>')
        if obj.can_edit_content:
            perms.append('<span style="color: blue;">ویرایش</span>')
        if obj.can_delete_content:
            perms.append('<span style="color: red;">حذف</span>')
        if obj.can_manage_users:
            perms.append('<span style="color: purple;">مدیریت کاربران</span>')
        if obj.can_view_analytics:
            perms.append('<span style="color: orange;">آمار</span>')
        if obj.can_access_admin:
            perms.append(
                '<span style="color: darkred; font-weight: bold;">ادمین</span>'
            )

        return (
            format_html(" | ".join(perms))
            if perms
            else format_html('<em style="color: #888;">بدون مجوز</em>')
        )

    permissions_summary.short_description = "مجوزها"

    def limits_summary(self, obj):
        """Show content limits summary"""
        limits = []
        if obj.max_posts_per_day:
            limits.append(f"{obj.max_posts_per_day} پست/روز")
        if obj.max_comments_per_day:
            limits.append(f"{obj.max_comments_per_day} نظر/روز")
        if obj.max_file_upload_size_mb:
            limits.append(f"{obj.max_file_upload_size_mb}MB فایل")

        return ", ".join(limits) if limits else "بدون محدودیت"

    limits_summary.short_description = "محدودیت‌ها"

    def users_count(self, obj):
        """Show number of users with this type"""
        count = getattr(obj, "users_count_num", obj.user_set.count())
        if count == 0:
            return format_html('<span style="color: #888;">0</span>')
        return format_html(f"<strong>{count}</strong>")

    users_count.short_description = "تعداد کاربران"

    def activate_user_types(self, request, queryset):
        """Activate selected user types"""
        updated = queryset.update(is_active=True)
        self.message_user(
            request, f"{updated} نوع کاربری فعال شد.", level=messages.SUCCESS
        )

    activate_user_types.short_description = "فعال کردن انواع کاربری انتخاب شده"

    def deactivate_user_types(self, request, queryset):
        """Deactivate selected user types"""
        updated = queryset.update(is_active=False)
        self.message_user(
            request, f"{updated} نوع کاربری غیرفعال شد.", level=messages.WARNING
        )

    deactivate_user_types.short_description = "غیرفعال کردن انواع کاربری انتخاب شده"

    def create_predefined_types(self, request, queryset):
        """Create predefined user types"""
        predefined_types = [
            {
                "name": "مدیر سیستم",
                "slug": "admin",
                "description": "دسترسی کامل به سیستم",
                "can_create_content": True,
                "can_edit_content": True,
                "can_delete_content": True,
                "can_manage_users": True,
                "can_view_analytics": True,
                "can_access_admin": True,
                "max_file_upload_size_mb": 100,
            },
            {
                "name": "مدیر محتوا",
                "slug": "manager",
                "description": "مدیریت محتوا و کاربران",
                "can_create_content": True,
                "can_edit_content": True,
                "can_delete_content": True,
                "can_manage_users": True,
                "can_view_analytics": True,
                "can_access_admin": True,
                "max_posts_per_day": 50,
                "max_comments_per_day": 100,
                "max_file_upload_size_mb": 50,
            },
            {
                "name": "ویرایشگر",
                "slug": "editor",
                "description": "ویرایش و مدیریت محتوا",
                "can_create_content": True,
                "can_edit_content": True,
                "can_delete_content": False,
                "can_manage_users": False,
                "can_view_analytics": True,
                "can_access_admin": True,
                "max_posts_per_day": 30,
                "max_comments_per_day": 50,
                "max_file_upload_size_mb": 25,
            },
            {
                "name": "نویسنده",
                "slug": "author",
                "description": "ایجاد و ویرایش محتوای شخصی",
                "can_create_content": True,
                "can_edit_content": True,
                "can_delete_content": False,
                "can_manage_users": False,
                "can_view_analytics": False,
                "can_access_admin": False,
                "max_posts_per_day": 10,
                "max_comments_per_day": 30,
                "max_file_upload_size_mb": 15,
            },
            {
                "name": "مشترک",
                "slug": "subscriber",
                "description": "کاربر عادی سیستم",
                "can_create_content": False,
                "can_edit_content": False,
                "can_delete_content": False,
                "can_manage_users": False,
                "can_view_analytics": False,
                "can_access_admin": False,
                "max_posts_per_day": 0,
                "max_comments_per_day": 10,
                "max_file_upload_size_mb": 5,
                "is_default": True,
            },
        ]

        created_count = 0
        for type_data in predefined_types:
            user_type, created = UserType.objects.get_or_create(
                slug=type_data["slug"], defaults=type_data
            )
            if created:
                created_count += 1

        if created_count > 0:
            self.message_user(
                request,
                f"{created_count} نوع کاربری پیش‌فرض ایجاد شد.",
                level=messages.SUCCESS,
            )
        else:
            self.message_user(
                request,
                "تمام انواع کاربری پیش‌فرض از قبل وجود دارند.",
                level=messages.INFO,
            )

    create_predefined_types.short_description = "ایجاد انواع کاربری پیش‌فرض"


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    # List display with user types
    list_display = [
        "username",
        "image_tag",
        "email",
        "mobile",
        "full_name_display",
        "user_type_display",
        "is_active",
        "is_staff",
        "email_status",
        "phone_status",
        "comments_count",
        "last_login",
        "created_at",
    ]

    list_editable = ["is_active", "is_staff"]

    list_filter = [
        "is_active",
        "is_staff",
        "is_superuser",
        UserTypeFilter,
        "is_phone_verified",
        "is_email_verified",
        HasCommentsFilter,
        EmailStatusFilter,
        ("created_at", JDateFieldListFilter),
        ("last_login", JDateFieldListFilter),
    ]

    search_fields = [
        "username",
        "email",
        "mobile",
        "first_name",
        "last_name",
        "slug",
        "comments__content",
        "user_type__name",
    ]

    # Enhanced actions - including export actions
    actions = [
        "send_email_action",
        "activate_users",
        "deactivate_users",
        "verify_emails",
        "verify_phones",
        "change_user_type_action",
        "promote_to_staff",
        "demote_from_staff",
        "export_to_excel",
        "export_to_pdf",
    ]

    # Pagination
    list_per_page = 50
    list_max_show_all = 200

    # Ordering
    ordering = ["-created_at"]

    # Add inlines
    inlines = [CommentInline]

    # Enhanced fieldsets for editing existing users
    fieldsets = (
        ("اطلاعات کاربری", {"fields": ("username", "email", "mobile", "slug")}),
        (
            "تصویر کاربر",
            {
                "fields": ("image",),
            },
        ),
        ("اطلاعات شخصی", {"fields": ("first_name", "last_name", "bio", "birth_date")}),
        (
            "نوع کاربری و دسترسی",
            {
                "fields": ("user_type", "is_active", "is_staff", "is_superuser"),
                "description": "نوع کاربری تعیین کننده سطح دسترسی کاربر است",
            },
        ),
        ("وضعیت تایید", {"fields": ("is_phone_verified", "is_email_verified")}),
        (
            "مجوزها",
            {"fields": ("groups", "user_permissions"), "classes": ("collapse",)},
        ),
        (
            "آمار فعالیت",
            {
                "fields": ("posts_count", "comments_count", "last_activity"),
                "classes": ("collapse",),
            },
        ),
        (
            "تاریخ‌ها",
            {
                "fields": ("last_login", "date_joined", "created_at"),
                "classes": ("collapse",),
            },
        ),
    )

    # Add fieldsets specifically for adding new users
    add_fieldsets = (
        (
            "اطلاعات کاربری ضروری",
            {
                "classes": ("wide",),
                "fields": ("username", "email", "mobile", "password1", "password2"),
            },
        ),
        (
            "اطلاعات شخصی",
            {
                "classes": ("wide",),
                "fields": ("first_name", "last_name"),
            },
        ),
        (
            "نوع کاربری و دسترسی",
            {
                "classes": ("wide",),
                "fields": ("user_type", "is_active", "is_staff"),
                "description": "نوع کاربری تعیین کننده سطح دسترسی کاربر است",
            },
        ),
        (
            "تصویر کاربر",
            {
                "classes": ("wide",),
                "fields": ("image",),
            },
        ),
    )

    readonly_fields = [
        "created_at",
        "image_preview",
        "date_joined",
        "last_login",
        "last_activity",
        "posts_count",
        "comments_count",
    ]

    def get_persian_font_path(self):
        """Get available Persian font path from system"""
        system = platform.system()

        if system == "Windows":
            paths = [
                "C:/Windows/Fonts/tahoma.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/calibri.ttf",
            ]
        elif system == "Linux":
            paths = [
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/TTF/arial.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            ]
        elif system == "Darwin":  # macOS
            paths = [
                "/System/Library/Fonts/Arial.ttf",
                "/Library/Fonts/Arial.ttf",
                "/System/Library/Fonts/Helvetica.ttc",
            ]
        else:
            return None

        for path in paths:
            if os.path.exists(path):
                return path
        return None

    def export_to_excel(self, request, queryset):
        import openpyxl

        """Export selected users to Persian Excel file"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "کاربران"

            # Set RTL direction for the worksheet
            ws.sheet_view.rightToLeft = True

            # Define headers in Persian
            headers = [
                "نام کاربری",
                "نام کامل",
                "ایمیل",
                "موبایل",
                "نوع کاربری",
                "وضعیت فعال",
                "عضو کادر",
                "ایمیل تایید شده",
                "تلفن تایید شده",
                "تعداد نظرات",
                "آخرین ورود",
                "تاریخ عضویت",
            ]

            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Write data
            for row, user in enumerate(queryset, 2):
                data = [
                    user.username or "",
                    user.get_full_name() or "بدون نام",
                    user.email or "",
                    user.mobile or "",
                    user.user_type.name if user.user_type else "پیش‌فرض",
                    "فعال" if user.is_active else "غیرفعال",
                    "بله" if user.is_staff else "خیر",
                    "تایید شده" if user.is_email_verified else "تایید نشده",
                    "تایید شده" if user.is_phone_verified else "تایید نشده",
                    user.comments.count(),
                    (
                        user.last_login.strftime("%Y/%m/%d %H:%M")
                        if user.last_login
                        else "هرگز"
                    ),
                    (
                        user.created_at.strftime("%Y/%m/%d %H:%M")
                        if user.created_at
                        else ""
                    ),
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Prepare response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.xlsx"

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Log success
            logger.info(f"✅ Excel export completed by admin {request.user.username}")
            logger.info(f"   📊 Exported {queryset.count()} users to {filename}")

            self.message_user(
                request,
                f"فایل اکسل شامل {queryset.count()} کاربر با موفقیت ایجاد شد.",
                level=messages.SUCCESS,
            )

            return response

        except Exception as e:
            logger.error(
                f"❌ Excel export failed by admin {request.user.username}: {str(e)}"
            )
            self.message_user(
                request, f"خطا در ایجاد فایل اکسل: {str(e)}", level=messages.ERROR
            )
            return redirect("admin:users_user_changelist")

    export_to_excel.short_description = "دریافت خروجی اکسل از کاربران انتخاب شده"

    def export_to_pdf(self, request, queryset):
        from reportlab.lib import colors

        """Export selected users to PDF with proper Persian BIDI support"""
        try:
            # Import BIDI support libraries
            try:
                from arabic_reshaper import reshape
                from bidi.algorithm import get_display

                bidi_available = True
            except ImportError:
                bidi_available = False
                logger.warning(
                    "BIDI libraries not available. Install: pip install python-bidi arabic-reshaper"
                )

            # Create response
            response = HttpResponse(content_type="application/pdf")
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_export_{timestamp}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Create PDF document
            doc = SimpleDocTemplate(
                response,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20,
            )

            # Font registration with better Persian support
            persian_font_available = False
            registered_font_name = "Helvetica"

            # Try to get Persian font
            font_path = self.get_persian_font_path()
            if font_path:
                try:
                    pdfmetrics.registerFont(TTFont("SystemFont", font_path))
                    registered_font_name = "SystemFont"
                    persian_font_available = True
                    logger.info(f"Successfully registered font from: {font_path}")
                except Exception as font_error:
                    logger.warning(f"Failed to register system font: {font_error}")

            # Helper function to process Persian text
            def process_persian_text(text):
                """Process Persian text for proper RTL display"""
                if not text or not isinstance(text, str):
                    return str(text) if text else ""

                if not bidi_available:
                    return text

                try:
                    # First reshape Arabic/Persian characters
                    reshaped_text = reshape(text)
                    # Then apply BIDI algorithm
                    bidi_text = get_display(reshaped_text)
                    return bidi_text
                except Exception as e:
                    logger.warning(f"Failed to process Persian text '{text}': {e}")
                    return text

            # Build content
            story = []
            styles = getSampleStyleSheet()

            # Create custom styles with RTL support
            if persian_font_available:
                title_style = ParagraphStyle(
                    "CustomTitle",
                    parent=styles["Title"],
                    fontName=registered_font_name,
                    fontSize=16,
                    alignment=1,  # Center
                    spaceAfter=20,
                    textColor=colors.HexColor("#366092"),
                    # Add RTL support if available
                    wordWrap="RTL" if bidi_available else "LTR",
                )

                info_style = ParagraphStyle(
                    "CustomInfo",
                    parent=styles["Normal"],
                    fontName=registered_font_name,
                    fontSize=11,
                    alignment=2,  # Right align for RTL
                    spaceAfter=6,
                    wordWrap="RTL" if bidi_available else "LTR",
                )
            else:
                title_style = styles["Title"]
                info_style = styles["Normal"]

            # Add title and metadata with proper Persian processing
            if persian_font_available:
                title_text = process_persian_text("گزارش کاربران سیستم")
                story.append(Paragraph(title_text, title_style))
                story.append(Spacer(1, 15))

                current_time = timezone.now().strftime("%Y/%m/%d - %H:%M")
                admin_name = request.user.get_full_name() or request.user.username

                # Process entire strings with BIDI, not just parts
                if bidi_available:
                    # Create complete Persian sentences first, then process
                    admin_line = f"{admin_name} :تهیه شده توسط"  # Reverse order for RTL
                    date_line = f"{current_time} :تاریخ تهیه"
                    count_line = f"{queryset.count()} :تعداد کاربران"

                    # Process complete lines
                    admin_line_processed = process_persian_text(admin_line)
                    date_line_processed = process_persian_text(date_line)
                    count_line_processed = process_persian_text(count_line)

                    story.append(Paragraph(admin_line_processed, info_style))
                    story.append(Paragraph(date_line_processed, info_style))
                    story.append(Paragraph(count_line_processed, info_style))
                else:
                    # Fallback without BIDI - use simple approach
                    story.append(Paragraph(f"تهیه شده توسط: {admin_name}", info_style))
                    story.append(Paragraph(f"تاریخ تهیه: {current_time}", info_style))
                    story.append(
                        Paragraph(f"تعداد کاربران: {queryset.count()}", info_style)
                    )
            else:
                # Fallback to English
                story.append(Paragraph("Users Export Report", title_style))
                story.append(Spacer(1, 15))

                current_time = timezone.now().strftime("%Y-%m-%d %H:%M")
                admin_name = request.user.get_full_name() or request.user.username

                story.append(Paragraph(f"Generated by: {admin_name}", info_style))
                story.append(Paragraph(f"Export date: {current_time}", info_style))
                story.append(Paragraph(f"Total users: {queryset.count()}", info_style))

            story.append(Spacer(1, 20))

            # Prepare table data with Persian processing
            table_data = []

            # Headers with proper Persian processing
            if persian_font_available and bidi_available:
                headers = [
                    process_persian_text("نام کاربری"),
                    process_persian_text("نام کامل"),
                    process_persian_text("ایمیل"),
                    process_persian_text("موبایل"),
                    process_persian_text("نوع کاربری"),
                    process_persian_text("وضعیت"),
                    process_persian_text("تاریخ عضویت"),
                ]
            elif persian_font_available:
                headers = [
                    "نام کاربری",
                    "نام کامل",
                    "ایمیل",
                    "موبایل",
                    "نوع کاربری",
                    "وضعیت",
                    "تاریخ عضویت",
                ]
            else:
                headers = [
                    "Username",
                    "Full Name",
                    "Email",
                    "Mobile",
                    "User Type",
                    "Status",
                    "Join Date",
                ]

            table_data.append(headers)

            # Data rows with proper Persian processing
            for user in queryset:
                if persian_font_available and bidi_available:
                    # Process all Persian text properly
                    username = user.username or "-"
                    full_name = process_persian_text(user.get_full_name() or "بدون نام")
                    email = user.email or "-"
                    mobile = user.mobile or "-"
                    user_type = process_persian_text(
                        user.user_type.name if user.user_type else "پیش‌فرض"
                    )
                    status = process_persian_text(
                        "فعال" if user.is_active else "غیرفعال"
                    )
                    join_date = (
                        user.created_at.strftime("%Y/%m/%d") if user.created_at else "-"
                    )

                    row = [
                        username,
                        full_name,
                        email,
                        mobile,
                        user_type,
                        status,
                        join_date,
                    ]

                elif persian_font_available:
                    # Without BIDI processing
                    row = [
                        user.username or "-",
                        user.get_full_name() or "بدون نام",
                        user.email or "-",
                        user.mobile or "-",
                        user.user_type.name if user.user_type else "پیش‌فرض",
                        "فعال" if user.is_active else "غیرفعال",
                        (
                            user.created_at.strftime("%Y/%m/%d")
                            if user.created_at
                            else "-"
                        ),
                    ]
                else:
                    # English fallback
                    full_name = user.get_full_name()
                    if full_name:
                        full_name = str(full_name)
                    else:
                        full_name = "No name"

                    user_type = (
                        str(user.user_type.name) if user.user_type else "Default"
                    )

                    row = [
                        user.username or "-",
                        full_name,
                        user.email or "-",
                        user.mobile or "-",
                        user_type,
                        "Active" if user.is_active else "Inactive",
                        (
                            user.created_at.strftime("%Y-%m-%d")
                            if user.created_at
                            else "-"
                        ),
                    ]

                table_data.append(row)

            # Create table with appropriate widths
            col_widths = [
                1.2 * inch,
                1.8 * inch,
                2.2 * inch,
                1.3 * inch,
                1.5 * inch,
                1 * inch,
                1.2 * inch,
            ]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Enhanced table styling for RTL
            table_style = [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    (
                        registered_font_name
                        if persian_font_available
                        else "Helvetica-Bold"
                    ),
                ),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                # Data rows styling
                (
                    "FONTNAME",
                    (0, 1),
                    (-1, -1),
                    registered_font_name if persian_font_available else "Helvetica",
                ),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                # Alignment for Persian columns (right-aligned)
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "RIGHT" if persian_font_available and bidi_available else "CENTER",
                ),
                # Grid and borders
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#366092")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]

            # Add alternating row colors
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table_style.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.lightgrey)
                    )
                else:
                    table_style.append(("BACKGROUND", (0, i), (-1, i), colors.white))

            table.setStyle(TableStyle(table_style))
            story.append(table)

            # Build PDF
            doc.build(story)

            # Enhanced logging
            logger.info(f"✅ PDF export completed by admin {request.user.username}")
            logger.info(f"   📊 Exported {queryset.count()} users to {filename}")
            logger.info(
                f"   🔤 Font used: {registered_font_name}, Persian support: {persian_font_available}"
            )
            logger.info(f"   🔄 BIDI support: {bidi_available}")

            success_msg = f"فایل PDF شامل {queryset.count()} کاربر با موفقیت ایجاد شد."
            if not persian_font_available:
                success_msg += " (از فونت پیش‌فرض استفاده شد)"
            elif not bidi_available:
                success_msg += " (بدون پشتیبانی کامل RTL)"

            self.message_user(request, success_msg, level=messages.SUCCESS)
            return response

        except Exception as e:
            logger.error(
                f"❌ PDF export failed by admin {request.user.username}: {str(e)}"
            )
            logger.exception("PDF export exception details:")
            self.message_user(
                request, f"خطا در ایجاد فایل PDF: {str(e)}", level=messages.ERROR
            )
            return redirect("admin:users_user_changelist")

        export_to_pdf.short_description = "دریافت خروجی PDF از کاربران انتخاب شده"

    def changelist_view(self, request, extra_context=None):
        """Override changelist view to add notifications and export buttons"""
        extra_context = extra_context or {}

        # Add your existing context
        extra_context["show_send_all_email_button"] = True
        extra_context["show_create_user_with_type_button"] = True
        extra_context["show_export_buttons"] = True
        extra_context["user_types"] = UserType.objects.filter(is_active=True)

        # Add admin message notifications
        if request.user.is_staff:
            try:
                unread_messages = AdminMessage.objects.filter(status="unread").order_by(
                    "-created_at"
                )[:3]
                if unread_messages.exists():
                    extra_context["admin_notifications"] = unread_messages
                    extra_context["notification_count"] = unread_messages.count()

                    # Add a Django message for immediate visibility
                    messages.info(
                        request,
                        f"📨 شما {unread_messages.count()} پیام خوانده نشده دارید.",
                        extra_tags="admin-notification",
                    )
            except Exception:
                pass

        return super().changelist_view(request, extra_context=extra_context)

    # Optimize queries
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        # Annotate with comment count for efficiency
        queryset = queryset.annotate(
            comments_count_num=Count("comments"),
            approved_comments_count=Count(
                "comments", filter=Q(comments__is_approved=True)
            ),
        )
        # Select related for efficiency
        queryset = queryset.select_related("user_type")
        return queryset

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height: 100px; border-radius: 5px;" />',
                obj.image.url,
            )
        return "بدون تصویر"

    image_preview.short_description = "پیش‌نمایش تصویر"

    def image_tag(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:40px; border-radius:4px;" />',
                obj.image.url,
            )
        return "-"

    image_tag.short_description = "تصویر"

    def full_name_display(self, obj):
        """Display full name or fallback"""
        full_name = obj.get_full_name().strip()
        if full_name:
            return full_name
        return format_html('<em style="color: #888;">بدون نام</em>')

    full_name_display.short_description = "نام کامل"

    def user_type_display(self, obj):
        """Display user type with color coding"""
        if not obj.user_type:
            return format_html('<em style="color: #888;">نامشخص</em>')

        # Color code based on permissions
        if obj.user_type.can_access_admin:
            color = "darkred"
            weight = "bold"
        elif obj.user_type.can_manage_users:
            color = "purple"
            weight = "bold"
        elif obj.user_type.can_create_content:
            color = "green"
            weight = "normal"
        else:
            color = "blue"
            weight = "normal"

        return format_html(
            '<span style="color: {}; font-weight: {};">{}</span>',
            color,
            weight,
            obj.user_type.name,
        )

    user_type_display.short_description = "نوع کاربری"

    def email_status(self, obj):
        """Display email validation status with colors"""
        if not obj.email:
            return format_html(
                '<span style="color: #888; font-style: italic;">بدون ایمیل</span>'
            )

        status_parts = []

        # Email verification status
        if obj.is_email_verified:
            status_parts.append('<span style="color: green;">✓ تایید شده</span>')
        else:
            status_parts.append('<span style="color: orange;">⚠ تایید نشده</span>')

        # Email validity
        try:
            EmailValidator.validate_email(obj.email)
            status_parts.append('<span style="color: green;">معتبر</span>')
        except ValidationError:
            status_parts.append('<span style="color: red;">نامعتبر</span>')

        return format_html(" | ".join(status_parts))

    email_status.short_description = "وضعیت ایمیل"

    def phone_status(self, obj):
        """Display phone verification status"""
        if not obj.mobile:
            return format_html(
                '<span style="color: #888; font-style: italic;">بدون شماره</span>'
            )

        if obj.is_phone_verified:
            return format_html('<span style="color: green;">✓ تایید شده</span>')
        else:
            return format_html('<span style="color: orange;">⚠ تایید نشده</span>')

    phone_status.short_description = "وضعیت تلفن"

    def comments_count(self, obj):
        """Display comment count with breakdown"""
        total = getattr(obj, "comments_count_num", obj.comments.count())
        approved = getattr(
            obj,
            "approved_comments_count",
            obj.comments.filter(is_approved=True).count(),
        )

        if total == 0:
            return format_html('<span style="color: #888;">بدون نظر</span>')

        pending = total - approved
        parts = [f"<strong>{total}</strong> کل"]

        if approved > 0:
            parts.append(f'<span style="color: green;">{approved} تایید</span>')

        if pending > 0:
            parts.append(f'<span style="color: orange;">{pending} انتظار</span>')

        return format_html(" | ".join(parts))

    comments_count.short_description = "نظرات"

    # Enhanced actions
    def send_email_action(self, request, queryset):
        selected = queryset.values_list("id", flat=True)
        selected_count = len(selected)

        logger.info(
            f"👨‍💼 Admin {request.user.username} initiated email action for {selected_count} users"
        )
        logger.debug(f"Selected user IDs: {list(selected)}")

        request.session["selected_users"] = list(selected)
        return HttpResponseRedirect(reverse("admin:send_email"))

    send_email_action.short_description = "ارسال ایمیل به کاربران انتخاب شده"

    def activate_users(self, request, queryset):
        """Activate selected users"""
        updated = queryset.update(is_active=True)
        logger.info(f"👨‍💼 Admin {request.user.username} activated {updated} users")
        self.message_user(request, f"{updated} کاربر فعال شد.", level=messages.SUCCESS)

    activate_users.short_description = "فعال کردن کاربران انتخاب شده"

    def deactivate_users(self, request, queryset):
        """Deactivate selected users"""
        updated = queryset.update(is_active=False)
        logger.info(f"👨‍💼 Admin {request.user.username} deactivated {updated} users")
        self.message_user(
            request, f"{updated} کاربر غیرفعال شد.", level=messages.WARNING
        )

    deactivate_users.short_description = "غیرفعال کردن کاربران انتخاب شده"

    def verify_emails(self, request, queryset):
        """Verify emails for selected users"""
        updated = (
            queryset.filter(email__isnull=False)
            .exclude(email="")
            .update(is_email_verified=True)
        )
        logger.info(
            f"👨‍💼 Admin {request.user.username} verified emails for {updated} users"
        )
        self.message_user(
            request, f"ایمیل {updated} کاربر تایید شد.", level=messages.SUCCESS
        )

    verify_emails.short_description = "تایید ایمیل کاربران انتخاب شده"

    def verify_phones(self, request, queryset):
        """Verify phones for selected users"""
        updated = (
            queryset.filter(mobile__isnull=False)
            .exclude(mobile="")
            .update(is_phone_verified=True)
        )
        logger.info(
            f"👨‍💼 Admin {request.user.username} verified phones for {updated} users"
        )
        self.message_user(
            request, f"شماره تلفن {updated} کاربر تایید شد.", level=messages.SUCCESS
        )

    verify_phones.short_description = "تایید شماره تلفن کاربران انتخاب شده"

    def change_user_type_action(self, request, queryset):
        """Change user type for selected users"""
        selected = queryset.values_list("id", flat=True)
        request.session["selected_users_for_type_change"] = list(selected)
        return HttpResponseRedirect(reverse("admin:change_user_type"))

    change_user_type_action.short_description = "تغییر نوع کاربری کاربران انتخاب شده"

    def promote_to_staff(self, request, queryset):
        """Promote users to staff status"""
        updated = queryset.update(is_staff=True)
        logger.info(
            f"👨‍💼 Admin {request.user.username} promoted {updated} users to staff"
        )
        self.message_user(
            request, f"{updated} کاربر به عضو کادر ارتقا یافت.", level=messages.SUCCESS
        )

    promote_to_staff.short_description = "ارتقا به عضو کادر"

    def demote_from_staff(self, request, queryset):
        """Demote users from staff status"""
        updated = queryset.update(is_staff=False)
        logger.info(
            f"👨‍💼 Admin {request.user.username} demoted {updated} users from staff"
        )
        self.message_user(
            request, f"{updated} کاربر از عضویت کادر خارج شد.", level=messages.WARNING
        )

    demote_from_staff.short_description = "خروج از عضویت کادر"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("send-email/", self.send_email_view, name="send_email"),
            path("send-email-all/", self.send_email_all_view, name="send_email_all"),
            path(
                "change-user-type/", self.change_user_type_view, name="change_user_type"
            ),
            path(
                "create-user-with-type/",
                self.create_user_with_type_view,
                name="create_user_with_type",
            ),
            # Export URLs
            path(
                "export-all-excel/", self.export_all_excel_view, name="export_all_excel"
            ),
            path("export-all-pdf/", self.export_all_pdf_view, name="export_all_pdf"),
        ]
        return custom_urls + urls

    def export_all_excel_view(self, request):
        """Export all users to Excel"""
        all_users = User.objects.all()
        return self.export_to_excel(request, all_users)

    def export_all_pdf_view(self, request):
        """Export all users to PDF"""
        all_users = User.objects.all()
        return self.export_to_pdf(request, all_users)

    def create_user_with_type_view(self, request):
        """Create user with specific type"""
        if request.method == "POST":
            return self._handle_create_user_post(request)
        else:
            return self._handle_create_user_get(request)

    def _handle_create_user_get(self, request):
        """Handle GET request for create user form"""
        user_types = UserType.objects.filter(is_active=True)
        context = {
            "title": "ایجاد کاربر با نوع مشخص",
            "user_types": user_types,
        }
        return render(request, "admin/create_user_with_type.html", context)

    def _handle_create_user_post(self, request):
        """Handle POST request for create user"""
        try:
            # Get form data
            username = request.POST.get("username")
            email = request.POST.get("email")
            mobile = request.POST.get("mobile")
            first_name = request.POST.get("first_name", "")
            last_name = request.POST.get("last_name", "")
            user_type_id = request.POST.get("user_type")
            password = request.POST.get("password")
            is_active = request.POST.get("is_active") == "on"
            is_staff = request.POST.get("is_staff") == "on"

            # Validation
            if not any([username, email, mobile]):
                messages.error(
                    request,
                    "حداقل یکی از فیلدهای نام کاربری، ایمیل یا موبایل باید پر شود.",
                )
                return self._handle_create_user_get(request)

            if not password:
                messages.error(request, "رمز عبور الزامی است.")
                return self._handle_create_user_get(request)

            # Get user type
            user_type = None
            if user_type_id:
                try:
                    user_type = UserType.objects.get(id=user_type_id, is_active=True)
                except UserType.DoesNotExist:
                    messages.error(request, "نوع کاربری انتخاب شده معتبر نیست.")
                    return self._handle_create_user_get(request)

            # Create user
            user = User.objects.create_user(
                username=username or None,
                email=email or None,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_active=is_active,
                is_staff=is_staff,
            )

            # Set additional fields
            if mobile:
                user.mobile = mobile
            if user_type:
                user.user_type = user_type

            user.save()

            logger.info(
                f"👨‍💼 Admin {request.user.username} created user: {user.get_display_name()} with type: {user_type.name if user_type else 'None'}"
            )

            messages.success(
                request,
                f'کاربر {user.get_display_name()} با نوع کاربری "{user_type.name if user_type else "پیش‌فرض"}" ایجاد شد.',
            )
            return redirect("admin:users_user_changelist")

        except Exception as e:
            logger.error(
                f"❌ Error creating user by admin {request.user.username}: {str(e)}"
            )
            messages.error(request, f"خطا در ایجاد کاربر: {str(e)}")
            return self._handle_create_user_get(request)

    def change_user_type_view(self, request):
        """Change user type for selected users"""
        if request.method == "POST":
            return self._handle_change_type_post(request)
        else:
            return self._handle_change_type_get(request)

    def _handle_change_type_get(self, request):
        """Handle GET request for change type form"""
        selected_users = request.session.get("selected_users_for_type_change", [])
        if not selected_users:
            messages.error(request, "هیچ کاربری انتخاب نشده است.")
            return redirect("admin:users_user_changelist")

        users = User.objects.filter(id__in=selected_users)
        user_types = UserType.objects.filter(is_active=True)

        context = {
            "title": "تغییر نوع کاربری",
            "users": users,
            "user_types": user_types,
            "total_users": len(users),
        }
        return render(request, "admin/change_user_type.html", context)

    def _handle_change_type_post(self, request):
        """Handle POST request for change type"""
        try:
            selected_users = request.session.get("selected_users_for_type_change", [])
            new_type_id = request.POST.get("user_type")

            if not selected_users:
                messages.error(request, "هیچ کاربری انتخاب نشده است.")
                return redirect("admin:users_user_changelist")

            # Get new user type
            new_type = None
            if new_type_id:
                try:
                    new_type = UserType.objects.get(id=new_type_id, is_active=True)
                except UserType.DoesNotExist:
                    messages.error(request, "نوع کاربری انتخاب شده معتبر نیست.")
                    return self._handle_change_type_get(request)

            # Update users
            users = User.objects.filter(id__in=selected_users)
            updated_count = 0

            for user in users:
                old_type = user.user_type
                user.user_type = new_type

                # Auto-update staff status based on new type
                if new_type and new_type.can_access_admin:
                    user.is_staff = True
                elif (
                    old_type
                    and old_type.can_access_admin
                    and new_type
                    and not new_type.can_access_admin
                ):
                    user.is_staff = False

                user.save()
                updated_count += 1

                logger.info(
                    f"👨‍💼 Admin {request.user.username} changed user type for {user.get_display_name()} from {old_type.name if old_type else 'None'} to {new_type.name if new_type else 'None'}"
                )

            # Clear session
            if "selected_users_for_type_change" in request.session:
                del request.session["selected_users_for_type_change"]

            type_name = new_type.name if new_type else "پیش‌فرض"
            messages.success(
                request, f'نوع کاربری {updated_count} کاربر به "{type_name}" تغییر کرد.'
            )
            return redirect("admin:users_user_changelist")

        except Exception as e:
            logger.error(
                f"❌ Error changing user types by admin {request.user.username}: {str(e)}"
            )
            messages.error(request, f"خطا در تغییر نوع کاربری: {str(e)}")
            return self._handle_change_type_get(request)

    # Email management methods
    def send_email_all_view(self, request):
        """Send email to all users view"""
        logger.info(
            f"👨‍💼 Admin {request.user.username} accessing send email to all users"
        )

        # Get all user IDs
        all_user_ids = list(User.objects.values_list("id", flat=True))
        request.session["selected_users"] = all_user_ids
        request.session["email_all_users"] = True

        logger.info(f"📊 Preparing to send email to all {len(all_user_ids)} users")

        return HttpResponseRedirect(reverse("admin:send_email"))

    def send_email_view(self, request):
        """Enhanced send email view with better error handling"""
        if request.method == "POST":
            return self._handle_email_post(request)
        else:
            return self._handle_email_get(request)

    def _handle_email_post(self, request):
        """Handle POST request for email sending"""
        form = EmailForm(request.POST)
        if not form.is_valid():
            logger.warning(f"⚠️ Invalid form submission by {request.user.username}")
            logger.debug(f"Form errors: {form.errors}")
            return self._render_email_form(request, form)

        try:
            # Get selected users
            selected_users = request.session.get("selected_users", [])
            is_email_all = request.session.get("email_all_users", False)

            if not selected_users:
                logger.warning(
                    f"⚠️ No users selected for email sending by {request.user.username}"
                )
                messages.error(request, "هیچ کاربری انتخاب نشده است.")
                return redirect("admin:users_user_changelist")

            logger.info(
                f"🎯 Processing email form submission by {request.user.username}"
            )
            logger.info(f"Selected users count: {len(selected_users)}")
            logger.info(f"Email all users: {is_email_all}")
            logger.info(f"Template: {form.cleaned_data['template'].name}")
            logger.info(f"Custom subject: {form.cleaned_data.get('subject', 'None')}")

            # Create and execute command
            command = SendEmailCommand(
                template_id=form.cleaned_data["template"].id,
                user_ids=selected_users,
                sender=request.user,
                custom_subject=form.cleaned_data.get("subject"),
                custom_content=form.cleaned_data.get("content"),
            )

            manager = EmailManager()
            manager.add_command(command)
            results = manager.execute_commands()

            # Process results
            self._process_email_results(request, results, is_email_all)

            # Clear session
            if "selected_users" in request.session:
                del request.session["selected_users"]
            if "email_all_users" in request.session:
                del request.session["email_all_users"]

            return redirect("admin:users_user_changelist")

        except Exception as e:
            error_msg = f"خطا در ارسال ایمیل: {str(e)}"
            messages.error(request, error_msg)
            logger.error(
                f"❌ Exception in admin email sending by {request.user.username}: {str(e)}"
            )
            logger.exception("Full exception details:")
            return self._render_email_form(request, form)

    def _handle_email_get(self, request):
        """Handle GET request for email form display"""
        form = EmailForm()
        logger.info(f"📋 Email form displayed to admin {request.user.username}")
        return self._render_email_form(request, form)

    def _render_email_form(self, request, form):
        """Render the email form with validation info"""
        selected_users = request.session.get("selected_users", [])
        is_email_all = request.session.get("email_all_users", False)
        users = User.objects.filter(id__in=selected_users)

        # Add validation info to context
        valid_users, invalid_users = EmailValidator.validate_users(users)

        logger.debug(
            f"Form display: {len(valid_users)} valid, {len(invalid_users)} invalid users"
        )

        context = {
            "form": form,
            "users": users,
            "valid_users": valid_users,
            "invalid_users": invalid_users,
            "is_email_all": is_email_all,
            "total_users": len(users),
            "title": "ارسال ایمیل به همه کاربران" if is_email_all else "ارسال ایمیل",
        }

        return render(request, "admin/send_email.html", context)

    def _process_email_results(self, request, results, is_email_all=False):
        """Process email sending results and display appropriate messages"""
        if results and results[0][0]:  # Check if successful
            success, message, details = results[0]

            # Create detailed success message
            total_users = details.get("total_users", 0)
            valid_users = details.get("valid_users", 0)
            invalid_users = details.get("invalid_users", 0)

            if is_email_all:
                if invalid_users > 0:
                    success_msg = f"ایمیل با موفقیت به {valid_users} کاربر از {total_users} کاربر موجود ارسال شد. ({invalid_users} کاربر نامعتبر نادیده گرفته شد)"
                    messages.success(request, success_msg)
                    self._add_invalid_user_warnings(request, details)
                else:
                    success_msg = (
                        f"ایمیل با موفقیت به تمام {valid_users} کاربر سیستم ارسال شد."
                    )
                    messages.success(request, success_msg)
            else:
                if invalid_users > 0:
                    success_msg = f"ایمیل با موفقیت به {valid_users} کاربر از {total_users} کاربر انتخاب شده ارسال شد. ({invalid_users} کاربر نامعتبر نادیده گرفته شد)"
                    messages.success(request, success_msg)
                    self._add_invalid_user_warnings(request, details)
                else:
                    success_msg = f"ایمیل با موفقیت به تمام {valid_users} کاربر انتخاب شده ارسال شد."
                    messages.success(request, success_msg)

            logger.info(
                f"✅ Email successfully processed by admin {request.user.username}"
            )
            logger.info(
                f"  📊 Results: {valid_users}/{total_users} users, {invalid_users} invalid"
            )

        else:
            error_msg = results[0][1] if results else "خطای نامشخص"
            messages.error(request, f"خطا در ارسال ایمیل: {error_msg}")
            logger.error(
                f"❌ Email sending failed by admin {request.user.username}: {error_msg}"
            )

    def _add_invalid_user_warnings(self, request, details):
        """Add warning messages for invalid users"""
        invalid_details = details.get("invalid_details", [])
        if invalid_details:
            warning_msg = "کاربران نامعتبر: "
            invalid_reasons = []
            for invalid_user in invalid_details:
                user = invalid_user["user"]
                issues = invalid_user["issues"]
                reason = []
                if "inactive_user" in issues:
                    reason.append("غیرفعال")
                if "invalid_email" in issues:
                    reason.append("ایمیل نامعتبر")
                invalid_reasons.append(f"{user.username} ({', '.join(reason)})")

            warning_msg += ", ".join(invalid_reasons)
            messages.warning(request, warning_msg)


@admin.register(Comment)
class CommentAdmin(admin.ModelAdmin):
    """Enhanced Comment admin"""

    list_display = ["user", "content_preview", "is_approved", "is_active", "created_at"]
    list_editable = ["is_approved", "is_active"]
    list_filter = ["is_approved", "is_active", ("created_at", JDateFieldListFilter)]
    search_fields = ["user__username", "user__email", "content"]
    ordering = ["-created_at"]
    list_per_page = 50

    actions = [
        "approve_comments",
        "reject_comments",
        "activate_comments",
        "deactivate_comments",
    ]

    def content_preview(self, obj):
        """Show preview of comment content"""
        content = obj.content[:100]
        if len(obj.content) > 100:
            content += "..."
        return format_html('<span title="{}">{}</span>', obj.content, content)

    content_preview.short_description = "متن نظر"

    def approve_comments(self, request, queryset):
        updated = queryset.update(is_approved=True)
        self.message_user(request, f"{updated} نظر تایید شد.", level=messages.SUCCESS)

    approve_comments.short_description = "تایید نظرات انتخاب شده"

    def reject_comments(self, request, queryset):
        updated = queryset.update(is_approved=False)
        self.message_user(request, f"{updated} نظر رد شد.", level=messages.WARNING)

    reject_comments.short_description = "رد نظرات انتخاب شده"

    def activate_comments(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"{updated} نظر فعال شد.", level=messages.SUCCESS)

    activate_comments.short_description = "فعال کردن نظرات انتخاب شده"

    def deactivate_comments(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"{updated} نظر غیرفعال شد.", level=messages.WARNING)

    deactivate_comments.short_description = "غیرفعال کردن نظرات انتخاب شده"


@admin.register(PasswordEntry)
class PasswordEntryAdmin(admin.ModelAdmin):
    list_display = ["user", "service_name", "username", "created_at", "updated_at"]
    list_filter = ["created_at", "updated_at"]
    search_fields = ["user__username", "service_name", "username"]
    readonly_fields = ["created_at", "updated_at"]
    ordering = ["-created_at"]

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(self.readonly_fields)
        if obj:  # Editing existing object
            readonly_fields.append("password")  # Make password readonly when editing
        return readonly_fields

    def has_change_permission(self, request, obj=None):
        has_permission = request.user.is_superuser
        if obj:
            password_logger.info(
                f"Admin change permission check - User: {request.user.username}, Object: {obj.service_name}, Permission: {has_permission}"
            )
        return has_permission

    def has_delete_permission(self, request, obj=None):
        has_permission = request.user.is_superuser
        if obj:
            password_logger.info(
                f"Admin delete permission check - User: {request.user.username}, Object: {obj.service_name}, Permission: {has_permission}"
            )
        return has_permission

    def has_view_permission(self, request, obj=None):
        has_permission = request.user.is_superuser
        if obj:
            password_logger.info(
                f"Admin view permission check - User: {request.user.username}, Object: {obj.service_name}, Permission: {has_permission}"
            )
        return has_permission

    def save_model(self, request, obj, form, change):
        if change:
            password_logger.info(
                f"Admin updating password entry - Admin: {request.user.username}, Entry: {obj.service_name}, User: {obj.user.username}"
            )
        else:
            password_logger.info(
                f"Admin creating password entry - Admin: {request.user.username}, Entry: {obj.service_name}, User: {obj.user.username}"
            )

        super().save_model(request, obj, form, change)

    def delete_model(self, request, obj):
        password_logger.info(
            f"Admin deleting password entry - Admin: {request.user.username}, Entry: {obj.service_name}, User: {obj.user.username}"
        )
        super().delete_model(request, obj)

    def changelist_view(self, request, extra_context=None):
        password_logger.info(
            f"Admin changelist accessed - Admin: {request.user.username}"
        )
        return super().changelist_view(request, extra_context)

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        if object_id:
            password_logger.info(
                f"Admin change form accessed - Admin: {request.user.username}, Object ID: {object_id}"
            )
        else:
            password_logger.info(
                f"Admin add form accessed - Admin: {request.user.username}"
            )
        return super().changeform_view(request, object_id, form_url, extra_context)


@admin.register(EmailTemplate)
class EmailTemplateAdmin(admin.ModelAdmin):
    list_display = ["name", "subject", "is_active", "created_at", "updated_at"]
    list_filter = ["is_active", ("created_at", JDateFieldListFilter)]
    search_fields = ["name", "subject"]
    readonly_fields = ["created_at", "updated_at"]

    # Add fieldsets for better organization
    fieldsets = (
        (None, {"fields": ("name", "subject", "content", "is_active")}),
        (
            "Timestamps",
            {"fields": ("created_at", "updated_at"), "classes": ("collapse",)},
        ),
    )

    def save_model(self, request, obj, form, change):
        if change:
            logger.info(
                f"📝 Email template updated: '{obj.name}' by {request.user.username}"
            )
        else:
            logger.info(
                f"➕ New email template created: '{obj.name}' by {request.user.username}"
            )
        super().save_model(request, obj, form, change)


class AdminMessageReadStatusInline(admin.TabularInline):
    """Inline for showing who has read the message"""

    model = AdminMessageReadStatus
    extra = 0
    readonly_fields = ["user", "read_at"]

    def has_add_permission(self, request, obj=None):
        return False


class AdminMessageReplyInline(admin.TabularInline):
    """Inline for message replies"""

    model = AdminMessageReply
    extra = 0
    readonly_fields = ["sender", "created_at"]
    fields = ["sender", "reply_text", "created_at"]


@admin.register(AdminMessage)
class AdminMessageAdmin(admin.ModelAdmin):
    """Admin interface for AdminMessage - Only visible to superusers"""

    list_display = [
        "subject",
        "sender_display",
        "priority_display",
        "status_display",
        "read_count",
        "created_at",
    ]

    list_filter = [
        "status",
        "priority",
        "sender",
        ("created_at", JDateFieldListFilter),
    ]

    search_fields = ["subject", "message", "sender__username", "sender__email"]

    readonly_fields = ["sender", "created_at", "read_at", "updated_at"]

    ordering = ["-created_at"]
    list_per_page = 25

    actions = ["mark_as_read", "mark_as_archived", "mark_as_unread"]

    inlines = [AdminMessageReplyInline, AdminMessageReadStatusInline]

    fieldsets = (
        ("اطلاعات پیام", {"fields": ("sender", "subject", "message", "priority")}),
        ("وضعیت", {"fields": ("status", "read_at")}),
        ("تاریخ‌ها", {"fields": ("created_at", "updated_at"), "classes": ("collapse",)}),
    )

    def get_queryset(self, request):
        """Add annotations for efficient queries"""
        return (
            super()
            .get_queryset(request)
            .annotate(read_count_num=Count("read_by", distinct=True))
            .select_related("sender")
        )

    def sender_display(self, obj):
        """Display sender with user type"""
        sender = obj.sender
        # Add safety check for sender existence
        if not sender:
            return format_html('<span style="color: #dc3545;">فرستنده حذف شده</span>')

        # Add safety check for user_type attribute
        try:
            type_name = (
                sender.get_user_type_display()
                if hasattr(sender, "user_type") and sender.user_type
                else "نامشخص"
            )
        except (AttributeError, ValueError):
            type_name = "نامشخص"

        # Add safety check for get_display_name method
        try:
            display_name = (
                sender.get_display_name()
                if hasattr(sender, "get_display_name")
                else str(sender)
            )
        except (AttributeError, ValueError):
            display_name = str(sender)

        return format_html(
            '<strong>{}</strong><br><small style="color: #666;">{}</small>',
            display_name,
            type_name,
        )

    sender_display.short_description = "فرستنده"

    def priority_display(self, obj):
        """Display priority with color and icon"""
        # Add safety checks for priority methods
        try:
            color_class = (
                obj.get_priority_color() if hasattr(obj, "get_priority_color") else ""
            )
            icon = obj.get_priority_icon() if hasattr(obj, "get_priority_icon") else ""
            priority_text = (
                obj.get_priority_display()
                if hasattr(obj, "get_priority_display")
                else str(obj.priority)
            )
        except (AttributeError, ValueError):
            color_class = ""
            icon = ""
            priority_text = str(obj.priority) if obj.priority else "نامشخص"

        return format_html(
            '<span class="{}">{} {}</span>', color_class, icon, priority_text
        )

    priority_display.short_description = "اولویت"

    def status_display(self, obj):
        """Display status with color coding"""
        colors = {
            "unread": "color: #d63384; font-weight: bold;",
            "read": "color: #198754;",
            "archived": "color: #6c757d;",
        }
        style = colors.get(obj.status, "")

        # Add safety check for get_status_display
        try:
            status_text = (
                obj.get_status_display()
                if hasattr(obj, "get_status_display")
                else str(obj.status)
            )
        except (AttributeError, ValueError):
            status_text = str(obj.status) if obj.status else "نامشخص"

        return format_html('<span style="{}">{}</span>', style, status_text)

    status_display.short_description = "وضعیت"

    def read_count(self, obj):
        """Show how many people have read the message"""
        try:
            count = getattr(obj, "read_count_num", None)
            if count is None:
                # Fallback to direct count if annotation failed
                count = obj.read_by.count() if hasattr(obj, "read_by") else 0
        except (AttributeError, ValueError):
            count = 0

        if count == 0:
            return format_html('<span style="color: #6c757d;">هیچ‌کس</span>')
        return format_html('<span style="color: #198754;">{} نفر</span>', count)

    read_count.short_description = "خوانده شده توسط"

    def mark_as_read(self, request, queryset):
        """Mark selected messages as read"""
        count = 0
        for message in queryset:
            try:
                if hasattr(message, "mark_as_read"):
                    message.mark_as_read(request.user)
                    count += 1
                else:
                    # Fallback method
                    message.status = "read"
                    message.read_at = timezone.now()
                    message.save()
                    count += 1
            except Exception as e:
                # Log error but continue with other messages
                self.message_user(
                    request,
                    f"خطا در علامت‌گذاری پیام {message.id}: {str(e)}",
                    messages.ERROR,
                )

        if count > 0:
            self.message_user(
                request,
                f"{count} پیام به عنوان خوانده شده علامت‌گذاری شد.",
                messages.SUCCESS,
            )

    mark_as_read.short_description = "علامت‌گذاری به عنوان خوانده شده"

    def mark_as_archived(self, request, queryset):
        """Archive selected messages"""
        try:
            updated = queryset.update(status="archived")
            self.message_user(request, f"{updated} پیام آرشیو شد.", messages.SUCCESS)
        except Exception as e:
            self.message_user(
                request, f"خطا در آرشیو کردن پیام‌ها: {str(e)}", messages.ERROR
            )

    mark_as_archived.short_description = "آرشیو کردن"

    def mark_as_unread(self, request, queryset):
        """Mark selected messages as unread"""
        try:
            updated = queryset.update(status="unread", read_at=None)
            self.message_user(
                request,
                f"{updated} پیام به عنوان خوانده نشده علامت‌گذاری شد.",
                messages.SUCCESS,
            )
        except Exception as e:
            self.message_user(
                request, f"خطا در علامت‌گذاری پیام‌ها: {str(e)}", messages.ERROR
            )

    mark_as_unread.short_description = "علامت‌گذاری به عنوان خوانده نشده"

    def has_module_permission(self, request):
        """Only superusers can access this module"""
        return request.user.is_superuser

    def has_view_permission(self, request, obj=None):
        """Only superusers can view messages"""
        return request.user.is_superuser

    def has_add_permission(self, request):
        """Superusers cannot add messages from admin (they come from message admins)"""
        return False

    def has_change_permission(self, request, obj=None):
        """Superusers can change status and add replies"""
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        """Only superusers can delete messages"""
        return request.user.is_superuser

    def changelist_view(self, request, extra_context=None):
        """Add notification count to changelist"""
        extra_context = extra_context or {}
        try:
            # Add safety check for get_unread_count method
            if hasattr(AdminMessage, "get_unread_count"):
                extra_context["unread_count"] = AdminMessage.get_unread_count()
            else:
                # Fallback to direct query
                extra_context["unread_count"] = AdminMessage.objects.filter(
                    status="unread"
                ).count()
        except Exception as e:
            # If there's an error, set count to 0
            extra_context["unread_count"] = 0

        return super().changelist_view(request, extra_context=extra_context)


@admin.register(AdminMessageReply)
class AdminMessageReplyAdmin(admin.ModelAdmin):
    """Admin for message replies"""

    list_display = ["original_message", "sender", "reply_preview", "created_at"]
    list_filter = [("created_at", JDateFieldListFilter), "sender"]
    search_fields = ["reply_text", "original_message__subject", "sender__username"]
    readonly_fields = ["created_at"]
    ordering = ["-created_at"]

    def reply_preview(self, obj):
        """Show preview of reply text"""
        if not obj.reply_text:
            return "پاسخ خالی"

        preview = obj.reply_text[:100]
        if len(obj.reply_text) > 100:
            preview += "..."
        return preview

    reply_preview.short_description = "پیش‌نمایش پاسخ"

    def has_module_permission(self, request):
        """Only superusers can access this module"""
        return request.user.is_superuser


# Customize admin site
admin.site.site_header = "پنل مدیریت"
admin.site.site_title = "پنل مدیریت"
admin.site.index_title = "خوش آمدید به پنل مدیریت"
